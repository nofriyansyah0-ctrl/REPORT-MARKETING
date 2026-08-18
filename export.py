"""
export.py
Menghasilkan laporan Excel (.xlsx) dari data yang tersimpan di database --
ringkasan, status terkini, leaderboard, dan riwayat live. Dipakai lewat
endpoint /api/export/excel di main.py.
"""

import io
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import database

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="FF7A1A", end_color="FF7A1A", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=16)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="666666")
BODY_FONT = Font(name="Arial", size=10)
LABEL_FONT = Font(name="Arial", bold=True, size=10)

WIB = timezone(timedelta(hours=7))


def _style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws, widths: dict):
    """widths: {kolom_index (1-based): lebar_karakter}"""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _fmt_wib(iso_string: str) -> str:
    if not iso_string:
        return "-"
    try:
        dt = datetime.fromisoformat(iso_string).astimezone(WIB)
        return dt.strftime("%d %b %Y, %H:%M WIB")
    except (ValueError, TypeError):
        return iso_string


def _region_label(region: str) -> str:
    return "Jogja" if region == "jogja" else "Luar Jogja"


def build_report(days: int = 7, region: str = "all") -> io.BytesIO:
    """
    Bangun workbook Excel lengkap. Return BytesIO siap dikirim sebagai
    file response (tidak ditulis ke disk).
    """
    wb = Workbook()
    now_wib = datetime.now(timezone.utc).astimezone(WIB)

    # ============== SHEET 1: RINGKASAN ==============
    ws1 = wb.active
    ws1.title = "Ringkasan"

    ws1["A1"] = "Laporan Live Monitor Akun TikTok"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Dibuat: {now_wib.strftime('%d %B %Y, %H:%M WIB')} · Rentang: {days} hari terakhir · Region: {_region_label(region) if region != 'all' else 'Semua'}"
    ws1["A2"].font = SUBTITLE_FONT

    all_status = database.get_all_status()
    if region != "all":
        all_status = [a for a in all_status if a.get("region") == region]

    total_accounts = len(all_status)
    live_now = len([a for a in all_status if a["is_live"]])
    leaderboard = database.get_leaderboard(days=days, region=region)
    total_sessions = sum(e["session_count"] for e in leaderboard)
    avg_viewers_overall = (
        round(sum(e["avg_viewers"] or 0 for e in leaderboard) / len(leaderboard))
        if leaderboard else 0
    )

    summary_rows = [
        ("Total akun dipantau", total_accounts),
        ("Sedang live sekarang", live_now),
        (f"Total sesi live ({days} hari terakhir)", total_sessions),
        ("Rata-rata penonton (semua akun aktif)", avg_viewers_overall),
    ]
    r = 4
    for label, value in summary_rows:
        ws1.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws1.cell(row=r, column=2, value=value).font = BODY_FONT
        r += 1

    if leaderboard:
        r += 1
        ws1.cell(row=r, column=1, value="Top 5 Akun Paling Aktif").font = LABEL_FONT
        r += 1
        headers = ["Akun", "Region", "Sesi Live", "Rata² Penonton", "Penonton Tertinggi"]
        for i, h in enumerate(headers, start=1):
            ws1.cell(row=r, column=i, value=h)
        _style_header_row(ws1, r, len(headers))
        r += 1
        for entry in leaderboard[:5]:
            ws1.cell(row=r, column=1, value=f"@{entry['username']}").font = BODY_FONT
            ws1.cell(row=r, column=2, value=_region_label(entry["region"])).font = BODY_FONT
            ws1.cell(row=r, column=3, value=entry["session_count"]).font = BODY_FONT
            ws1.cell(row=r, column=4, value=entry["avg_viewers"] or "-").font = BODY_FONT
            ws1.cell(row=r, column=5, value=entry["max_viewers"] or "-").font = BODY_FONT
            r += 1

    _auto_width(ws1, {1: 32, 2: 14, 3: 14, 4: 16, 5: 18})

    # ============== SHEET 2: STATUS SAAT INI ==============
    ws2 = wb.create_sheet("Status Saat Ini")
    headers2 = ["Username", "Region", "Status", "Judul Live", "Jumlah Penonton", "Terakhir Dicek"]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=i, value=h)
    _style_header_row(ws2, 1, len(headers2))

    sorted_status = sorted(all_status, key=lambda a: (not a["is_live"], a["username"]))
    for row_idx, acc in enumerate(sorted_status, start=2):
        ws2.cell(row=row_idx, column=1, value=f"@{acc['username']}").font = BODY_FONT
        ws2.cell(row=row_idx, column=2, value=_region_label(acc.get("region", "jogja"))).font = BODY_FONT
        ws2.cell(row=row_idx, column=3, value="LIVE" if acc["is_live"] else "Offline").font = BODY_FONT
        ws2.cell(row=row_idx, column=4, value=acc.get("title") or "-").font = BODY_FONT
        ws2.cell(row=row_idx, column=5, value=acc.get("viewer_count") if acc.get("viewer_count") is not None else "-").font = BODY_FONT
        ws2.cell(row=row_idx, column=6, value=_fmt_wib(acc.get("last_checked"))).font = BODY_FONT
    ws2.freeze_panes = "A2"
    _auto_width(ws2, {1: 26, 2: 12, 3: 10, 4: 34, 5: 16, 6: 22})

    # ============== SHEET 3: LEADERBOARD ==============
    ws3 = wb.create_sheet(f"Leaderboard ({days} Hari)")
    headers3 = ["Peringkat", "Username", "Region", "Sesi Live", "Rata² Penonton", "Penonton Tertinggi", "Terakhir Live"]
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=i, value=h)
    _style_header_row(ws3, 1, len(headers3))

    for row_idx, entry in enumerate(leaderboard, start=2):
        rank = row_idx - 1
        ws3.cell(row=row_idx, column=1, value=rank).font = BODY_FONT
        ws3.cell(row=row_idx, column=2, value=f"@{entry['username']}").font = BODY_FONT
        ws3.cell(row=row_idx, column=3, value=_region_label(entry["region"])).font = BODY_FONT
        ws3.cell(row=row_idx, column=4, value=entry["session_count"]).font = BODY_FONT
        ws3.cell(row=row_idx, column=5, value=entry["avg_viewers"] or "-").font = BODY_FONT
        ws3.cell(row=row_idx, column=6, value=entry["max_viewers"] or "-").font = BODY_FONT
        ws3.cell(row=row_idx, column=7, value=_fmt_wib(entry["last_live_at"])).font = BODY_FONT
    ws3.freeze_panes = "A2"
    _auto_width(ws3, {1: 10, 2: 26, 3: 12, 4: 12, 5: 16, 6: 18, 7: 22})

    # ============== SHEET 4: RIWAYAT LIVE ==============
    ws4 = wb.create_sheet("Riwayat Live")
    headers4 = ["Username", "Judul Live", "Mulai", "Selesai", "Durasi (menit)", "Peak Penonton"]
    for i, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=i, value=h)
    _style_header_row(ws4, 1, len(headers4))

    history = database.get_history(limit=500)
    # Filter berdasarkan region kalau diminta (join manual lewat all_status)
    region_map = {a["username"]: a.get("region", "jogja") for a in database.get_all_status()}
    if region != "all":
        history = [h for h in history if region_map.get(h["username"]) == region]

    for row_idx, item in enumerate(history, start=2):
        duration_min = "-"
        if item["started_at"]:
            try:
                start = datetime.fromisoformat(item["started_at"])
                end = datetime.fromisoformat(item["ended_at"]) if item["ended_at"] else datetime.now(timezone.utc)
                duration_min = round((end - start).total_seconds() / 60)
            except (ValueError, TypeError):
                pass
        ws4.cell(row=row_idx, column=1, value=f"@{item['username']}").font = BODY_FONT
        ws4.cell(row=row_idx, column=2, value=item.get("title") or "-").font = BODY_FONT
        ws4.cell(row=row_idx, column=3, value=_fmt_wib(item["started_at"])).font = BODY_FONT
        ws4.cell(row=row_idx, column=4, value=_fmt_wib(item["ended_at"]) if item["ended_at"] else "Masih live").font = BODY_FONT
        ws4.cell(row=row_idx, column=5, value=duration_min).font = BODY_FONT
        ws4.cell(row=row_idx, column=6, value=item.get("peak_viewer_count") or "-").font = BODY_FONT
    ws4.freeze_panes = "A2"
    _auto_width(ws4, {1: 26, 2: 34, 3: 22, 4: 22, 5: 16, 6: 16})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
